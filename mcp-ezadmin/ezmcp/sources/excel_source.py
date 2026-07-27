"""엑셀·CSV 소스 (design/02) — 읽기 전용.

`data/orders|inventory|returns` 폴더를 재귀 스캔해 표준 레코드로 만든다.
파일명은 자유이며, 같은 키가 겹치면 더 최신 파일의 행이 이긴다.
"""

from __future__ import annotations

import csv
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from .. import normalize as nz
from ..dates import fmt_dt, kst, now_kst
from .base import KIND_LABEL, Source, SourceResult

log = logging.getLogger(__name__)

EXCEL_EXTS = {".xlsx", ".xlsm"}
CSV_EXTS = {".csv"}
LEGACY_EXTS = {".xls"}
MAX_ROWS_PER_FILE = 200_000

_cache: dict[tuple, SourceResult] = {}


def _mtime_kst(path: Path) -> datetime:
    """파일 수정시각을 KST 벽시계 값으로 (PC의 시스템 시간대와 무관하게 일관되게)."""
    return datetime.fromtimestamp(path.stat().st_mtime, tz=kst()).replace(tzinfo=None)


def _list_files(folder: Path) -> tuple[list[Path], list[str]]:
    """폴더 재귀 스캔 → (읽을 파일 목록, 경고)."""
    warnings: list[str] = []
    if not folder.is_dir():
        return [], warnings

    files: list[Path] = []
    for path in sorted(folder.rglob("*")):
        if not path.is_file() or path.name.startswith("~$"):
            continue
        ext = path.suffix.lower()
        if ext in EXCEL_EXTS or ext in CSV_EXTS:
            files.append(path)
        elif ext in LEGACY_EXTS:
            warnings.append(
                "%s 는 구형 엑셀(.xls)이라 읽지 못했습니다. "
                "xlsx 로 다시 내려받거나 '다른 이름으로 저장 → xlsx' 해 주세요." % path.name
            )
    files.sort(key=lambda p: (p.stat().st_mtime, p.name))
    return files, warnings


def _signature(files: list[Path]) -> tuple:
    sig = []
    for path in files:
        try:
            st = path.stat()
            sig.append((str(path), int(st.st_mtime), st.st_size))
        except OSError:
            sig.append((str(path), 0, 0))
    return tuple(sig)


def _read_xlsx(path: Path) -> tuple[list[list[Any]], str]:
    """xlsx → 행 목록. 읽기 전용 모드로 열고 반드시 닫는다."""
    from openpyxl import load_workbook

    wb = None
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        for sheet in wb.worksheets:
            rows: list[list[Any]] = []
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i >= MAX_ROWS_PER_FILE:
                    break
                rows.append(list(row))
            if any(any(c is not None and str(c).strip() for c in r) for r in rows):
                return rows, ""
        return [], ""
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:  # noqa: BLE001
                pass


def _read_csv(path: Path) -> tuple[list[list[Any]], str]:
    """CSV → 행 목록. utf-8-sig 실패 시 cp949."""
    last_error = ""
    for encoding in ("utf-8-sig", "cp949"):
        try:
            with path.open("r", encoding=encoding, newline="") as fh:
                return [list(row) for row in csv.reader(fh)][:MAX_ROWS_PER_FILE], ""
        except UnicodeDecodeError as exc:
            last_error = str(exc)
            continue
    return [], "%s 의 문자 인코딩을 알 수 없어 건너뛰었습니다 (utf-8/cp949 모두 실패)." % path.name


def _read_rows(path: Path) -> tuple[list[list[Any]], str]:
    ext = path.suffix.lower()
    try:
        if ext in CSV_EXTS:
            return _read_csv(path)
        return _read_xlsx(path)
    except PermissionError:
        return [], (
            "%s 파일이 열려 있어 읽지 못했습니다. 엑셀에서 파일을 닫고 다시 시도해 주세요."
            % path.name
        )
    except Exception as exc:  # noqa: BLE001 - 파일 하나 때문에 전체가 죽지 않게
        log.warning("파일 읽기 실패 %s: %s", path.name, exc)
        return [], "%s 를 읽지 못했습니다 (%s)." % (path.name, type(exc).__name__)


class ExcelSource(Source):
    name = "excel"

    def __init__(self, cfg):
        self.cfg = cfg

    def _cfg_fingerprint(self) -> str:
        return json.dumps(
            [self.cfg.columns, self.cfg.status_keywords], sort_keys=True, ensure_ascii=False
        )

    def fetch(self, kind: str, **kwargs) -> SourceResult:
        folder = self.cfg.data_subdir(kind)
        files, warnings = _list_files(folder)
        cache_key = (str(folder), _signature(files), self._cfg_fingerprint(), kind)
        cached = _cache.get(cache_key)
        if cached is not None:
            return cached

        result = self._parse(kind, folder, files, warnings)
        _cache[cache_key] = result
        return result

    def _parse(
        self, kind: str, folder: Path, files: list[Path], warnings: list[str]
    ) -> SourceResult:
        keywords = nz.status_keywords(self.cfg.status_keywords)
        merged: dict[tuple, dict[str, Any]] = {}
        file_infos: list[dict[str, str]] = []
        latest: datetime | None = None

        for path in files:
            mtime = _mtime_kst(path)
            rows, error = _read_rows(path)
            if error:
                warnings.append(error)
                continue
            if not rows:
                warnings.append("%s 에 읽을 데이터가 없습니다." % path.name)
                continue

            header_idx = nz.detect_header_row(rows, kind, self.cfg.columns)
            if header_idx is None:
                warnings.append(
                    "%s 에서 %s 데이터의 머리글(주문번호·상품명 등)을 찾지 못해 건너뛰었습니다."
                    % (path.name, KIND_LABEL.get(kind, kind))
                )
                continue

            header_cells = list(rows[header_idx])
            mapping = nz.map_columns(header_cells, kind, self.cfg.columns)
            count = 0
            for row in rows[header_idx + 1 :]:
                rec = nz.build_record(kind, row, mapping, header_cells, keywords)
                if rec is None:
                    continue
                rec["_file"] = path.name
                rec["_account"] = ""
                merged[nz.dedupe_key(kind, rec)] = rec
                count += 1

            file_infos.append(
                {"파일": path.name, "수정시각": fmt_dt(mtime), "행수": count}
            )
            latest = mtime if latest is None or mtime > latest else latest

        if latest is not None:
            age_hours = (now_kst() - latest).total_seconds() / 3600
            if age_hours > self.cfg.stale_hours:
                warnings.append(
                    "%s 데이터가 %d시간 전 파일입니다. 이지어드민에서 최신 파일을 내려받아 주세요."
                    % (KIND_LABEL.get(kind, kind), int(age_hours))
                )

        return SourceResult(
            records=list(merged.values()),
            files=file_infos,
            warnings=warnings,
            latest=latest,
            source="excel",
            ok=True,
        )


def clear_cache() -> None:
    _cache.clear()
