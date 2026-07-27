"""이지어드민풍 샘플 데이터 생성 (design/08).

실제 고객 정보는 절대 쓰지 않는다 — 전부 가상 이름·전화·주소다.
실데이터를 넣기 전에 서버 동작을 확인하거나 selftest 를 돌릴 때 사용한다.

사용법:
    python scripts/make_sample_data.py            # data/ 에 생성
    python scripts/make_sample_data.py --out 경로  # 다른 폴더에 생성
"""

from __future__ import annotations

import argparse
import csv
import os
import random
import sys
from datetime import datetime, time, timedelta
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

# 한국 Windows에서 출력을 파일로 리다이렉트할 때의 cp949 인코딩 오류 방지
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from openpyxl import Workbook  # noqa: E402

from ezmcp.dates import now_kst  # noqa: E402

SEED = 20260727

ORDER_HEADERS = [
    "주문번호",
    "판매처",
    "주문일시",
    "주문상태",
    "배송상태",
    "상품코드",
    "상품명",
    "옵션",
    "수량",
    "결제금액",
    "수취인",
    "주문자",
    "휴대폰",
    "주소",
    "택배사",
    "송장번호",
    "발송일",
    "배송메시지",
]

INVENTORY_HEADERS = [
    "상품코드",
    "상품명",
    "옵션",
    "브랜드",
    "현재고",
    "가용재고",
    "안전재고",
    "로케이션",
]

RETURN_HEADERS = [
    "주문번호",
    "상품명",
    "옵션",
    "수량",
    "접수일",
    "클레임구분",
    "처리상태",
    "사유",
    "수거택배사",
    "수거송장번호",
]

NAMES = [
    "김하늘",
    "이도윤",
    "박서준",
    "최민서",
    "정예린",
    "강태호",
    "윤지우",
    "임채원",
    "오세훈",
    "한소미",
    "남궁민수",
    "서지훈",
]
ADDRESSES = [
    "서울 강남구 역삼동 123-4 5층",
    "경기 성남시 분당구 정자동 77 101동 1203호",
    "부산 해운대구 우동 55-1",
    "인천 연수구 송도동 12-3 오피스텔 908호",
    "대구 수성구 범어동 8-8 상가 2층",
]
PRODUCTS = {
    "LM": [
        ("케이블 니트 가디건", ["아이보리/S", "아이보리/M", "블랙/M"]),
        ("실크 블라우스", ["화이트/S", "화이트/M"]),
        ("울 코트", ["카멜/M", "카멜/L"]),
        ("플리츠 스커트", ["네이비/S", "네이비/M"]),
    ],
    "VD": [
        ("오버핏 셔츠", ["블루/L", "블루/XL"]),
        ("데님 팬츠", ["인디고/28", "인디고/30", "인디고/32"]),
        ("레더 자켓", ["블랙/M", "블랙/L"]),
        ("니트 베스트", ["그레이/F"]),
    ],
}
MALLS = {
    "LM": ["루미네 자사몰", "스마트스토어-루미네", "쿠팡"],
    "VD": ["베르디 자사몰", "쿠팡"],
}
COURIERS = ["CJ대한통운", "한진택배", "롯데택배"]
MEMOS = ["부재 시 경비실", "안전하게 부탁드려요", "", "문 앞에 놔주세요"]

SHIPPED_STATUSES = ("배송중", "배송완료", "구매확정")
NO_TRACKING_STATUSES = (
    "신규주문",
    "결제완료",
    "배송준비중",
    "상품준비중",
    "보류",
    "주문취소",
    "발주대기",
)


def _fmt(dt: datetime | None) -> str:
    return dt.strftime("%Y-%m-%d %H:%M") if dt else ""


def _make_order_rows(rng: random.Random) -> list[list]:
    now = now_kst()
    today = now.date()
    rows: list[list] = []
    seq = 0

    def build(order_dt: datetime, status: str, ship_status: str = "") -> list:
        nonlocal seq
        seq += 1
        prefix = rng.choice(["LM", "VD"])
        name, options = rng.choice(PRODUCTS[prefix])
        option = rng.choice(options)
        code = "%s-%03d" % (prefix, rng.randint(1, 40))
        qty = rng.randint(1, 3)
        price = rng.choice([39000, 59000, 89000, 129000, 189000])
        tracked = status in SHIPPED_STATUSES
        courier = rng.choice(COURIERS) if tracked else ""
        tracking = "%d" % rng.randint(100000000000, 999999999999) if tracked else ""
        shipped = order_dt + timedelta(days=1) if tracked else None
        return [
            "%s-%04d" % (order_dt.strftime("%Y%m%d"), seq),
            rng.choice(MALLS[prefix]),
            _fmt(order_dt),
            status,
            ship_status,
            code,
            "[%s] %s" % ("루미네" if prefix == "LM" else "베르디", name),
            option,
            qty,
            qty * price,
            rng.choice(NAMES),
            rng.choice(NAMES),
            "010-%04d-%04d" % (rng.randint(1000, 9999), rng.randint(1000, 9999)),
            rng.choice(ADDRESSES),
            courier,
            tracking,
            _fmt(shipped),
            rng.choice(MEMOS),
        ]

    # 기간 경계 검증용 2건
    boundary_today = build(datetime.combine(today, time.min), "신규주문")
    boundary_today[0] = "BOUNDARY-TODAY-0000"
    rows.append(boundary_today)

    boundary_yesterday = build(
        datetime.combine(today - timedelta(days=1), time(23, 59)), "신규주문"
    )
    boundary_yesterday[0] = "BOUNDARY-YDAY-2359"
    rows.append(boundary_yesterday)

    # 미매핑 상태값(기타 버킷) 3건 — data_status 의 '기타상태값' 검증용
    for _ in range(3):
        rows.append(build(datetime.combine(today - timedelta(days=5), time(10, 30)), "발주대기"))

    # 오래된 미출고 3건
    for _ in range(3):
        rows.append(
            build(datetime.combine(today - timedelta(days=5), time(14, 0)), "배송준비중")
        )

    # 최근 출고 건 (택배사 3종 고루 섞기)
    for i in range(8):
        order_dt = datetime.combine(
            today - timedelta(days=(i % 5) + 1), time(9 + (i % 8), 15)
        )
        row = build(order_dt, "배송중" if i % 2 else "배송완료", "배송중" if i % 2 else "배송완료")
        row[14] = COURIERS[i % 3]
        rows.append(row)

    # 반품·교환 각 2건
    for status in ("반품접수", "반품접수", "교환접수", "교환접수"):
        rows.append(
            build(datetime.combine(today - timedelta(days=rng.randint(2, 6)), time(11, 0)), status)
        )

    # 나머지는 최근 14일에 분포
    remaining = 60 - len(rows)
    pool = [
        "신규주문",
        "결제완료",
        "배송준비중",
        "상품준비중",
        "배송중",
        "배송완료",
        "구매확정",
        "주문취소",
        "보류",
    ]
    for _ in range(remaining):
        days_ago = rng.randint(0, 13)
        order_dt = datetime.combine(
            today - timedelta(days=days_ago), time(rng.randint(9, 20), rng.choice([5, 20, 35, 50]))
        )
        status = rng.choice(pool)
        ship = status if status in SHIPPED_STATUSES else ""
        rows.append(build(order_dt, status, ship))

    return rows


def write_orders(folder: Path, rng: random.Random) -> list[list]:
    folder.mkdir(parents=True, exist_ok=True)
    rows = _make_order_rows(rng)

    wb = Workbook()
    ws = wb.active
    ws.title = "주문목록"
    ws.append(["이지어드민 주문 목록 (조회조건: 최근 14일)"])  # 헤더 위 제목행 — 탐지 검증용
    ws.append(ORDER_HEADERS)
    for row in rows:
        ws.append(row)
    wb.save(folder / "주문_샘플.xlsx")
    wb.close()
    return rows


def write_orders_csv(folder: Path, xlsx_rows: list[list], rng: random.Random) -> None:
    """cp949 CSV 10행. 그중 3행은 xlsx 와 같은 키로 중복 제거를 검증한다."""
    dup_sources = xlsx_rows[20:23]
    rows: list[list] = []

    for source in dup_sources:
        row = list(source)
        row[3] = "배송완료"  # 나중 파일이 이기는지 확인용
        row[4] = "배송완료"
        row[14] = "CJ대한통운"
        row[15] = "%d" % rng.randint(100000000000, 999999999999)
        rows.append(row)

    now = now_kst()
    seq = 900
    for _ in range(7):
        seq += 1
        prefix = rng.choice(["LM", "VD"])
        name, options = rng.choice(PRODUCTS[prefix])
        order_dt = now - timedelta(days=rng.randint(0, 6), hours=rng.randint(1, 9))
        rows.append(
            [
                "%s-%04d" % (order_dt.strftime("%Y%m%d"), seq),
                rng.choice(MALLS[prefix]),
                _fmt(order_dt),
                rng.choice(["신규주문", "배송준비중", "배송완료"]),
                "",
                "%s-%03d" % (prefix, rng.randint(41, 60)),
                "[%s] %s" % ("루미네" if prefix == "LM" else "베르디", name),
                rng.choice(options),
                1,
                59000,
                rng.choice(NAMES),
                rng.choice(NAMES),
                "010-%04d-%04d" % (rng.randint(1000, 9999), rng.randint(1000, 9999)),
                rng.choice(ADDRESSES),
                "",
                "",
                "",
                "",
            ]
        )

    path = folder / "주문_샘플2.csv"
    with path.open("w", encoding="cp949", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(ORDER_HEADERS)
        writer.writerows(rows)

    # xlsx 보다 확실히 최신이 되게 해서 '나중 파일이 이긴다' 규칙을 검증한다
    stamp = (folder / "주문_샘플.xlsx").stat().st_mtime + 10
    os.utime(path, (stamp, stamp))


def write_inventory(folder: Path, rng: random.Random) -> None:
    folder.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "현재고"
    ws.append(INVENTORY_HEADERS)

    rows = []
    for i in range(30):
        prefix = "LM" if i % 2 == 0 else "VD"
        name, options = rng.choice(PRODUCTS[prefix])
        low = i < 5  # 재고부족 5건
        available = 1 if low else rng.randint(20, 80)
        rows.append(
            [
                "%s-%03d" % (prefix, i + 1),
                "[%s] %s" % ("루미네" if prefix == "LM" else "베르디", name),
                rng.choice(options),
                "루미네" if prefix == "LM" else "베르디",
                available + rng.randint(0, 5),
                available,
                5,
                "%s-%02d-%02d" % (chr(ord("A") + i % 4), (i % 9) + 1, (i % 6) + 1),
            ]
        )
    for row in rows:
        ws.append(row)
    wb.save(folder / "재고_샘플.xlsx")
    wb.close()


def write_returns(folder: Path, order_rows: list[list], rng: random.Random) -> None:
    folder.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "클레임"
    ws.append(RETURN_HEADERS)

    claims = [r for r in order_rows if r[3] in ("반품접수", "교환접수")]
    while len(claims) < 8:
        claims.append(rng.choice(order_rows))

    today = now_kst().date()
    for i, source in enumerate(claims[:8]):
        with_tracking = i < 4
        ws.append(
            [
                source[0],
                source[6],
                source[7],
                1,
                _fmt(datetime.combine(today - timedelta(days=i), time(13, 0))),
                "교환" if "교환" in str(source[3]) else "반품",
                "수거중" if with_tracking else "접수",
                rng.choice(["단순변심", "사이즈 교환", "제품 불량"]),
                COURIERS[i % 3] if with_tracking else "",
                "%d" % rng.randint(100000000000, 999999999999) if with_tracking else "",
            ]
        )
    wb.save(folder / "반품_샘플.xlsx")
    wb.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="이지어드민풍 샘플 데이터 생성 (가상 데이터)")
    parser.add_argument("--out", default="", help="생성 위치 (기본: 프로젝트의 data 폴더)")
    args = parser.parse_args()

    base = Path(args.out).expanduser() if args.out else ROOT / "data"
    rng = random.Random(SEED)

    order_rows = write_orders(base / "orders", rng)
    write_orders_csv(base / "orders", order_rows, rng)
    write_inventory(base / "inventory", rng)
    write_returns(base / "returns", order_rows, rng)

    print("샘플 데이터를 만들었습니다: %s" % base)
    print("  orders/주문_샘플.xlsx (60행) + 주문_샘플2.csv (10행, cp949, 3행 중복)")
    print("  inventory/재고_샘플.xlsx (30행, 재고부족 5행)")
    print("  returns/반품_샘플.xlsx (8행, 수거송장 4행)")
    print("※ 전부 가상 데이터입니다. 실데이터를 넣기 전 테스트용으로만 쓰세요.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
