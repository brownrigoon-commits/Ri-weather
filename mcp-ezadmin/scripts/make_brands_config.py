"""주문 엑셀의 '공급처'·'판매처' 실제 값에서 config.json 의 brands 섹션을 만든다.

config.json 은 .gitignore 대상이라 PC 간에 옮겨지지 않는다. 다른 PC에서도 브랜드 설정을
손으로 다시 적지 않도록, 실제 데이터에서 그대로 생성한다.

- 브랜드 키: 공급처 값에서 접두어(`자체제작_` 등)를 뗀 이름
- aliases: 공급처 원문 값 전부. brand_col 매칭은 정확일치라 원문이 반드시 들어가야 한다
- malls: 브랜드명이 들어간 판매처 문자열. 공급처가 빈 행을 보완한다

사용법:
    python scripts/make_brands_config.py data/orders
    python scripts/make_brands_config.py data/orders --config config.json
    python scripts/make_brands_config.py data/orders --dry-run   # 파일을 쓰지 않고 결과만 출력
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

# 한국 Windows에서 출력을 파일로 리다이렉트할 때의 cp949 인코딩 오류 방지
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from openpyxl import load_workbook  # noqa: E402

# 공급처 값 앞에 붙는, 브랜드명이 아닌 접두어
PREFIXES = ("자체제작_", "vinicunca.co_", "EVOKE ERA 이보크에라_", "29cm_")

# 같은 브랜드의 영문·표기 변형. 판매처 문자열 매칭에도 함께 쓰인다.
EXTRA_ALIASES: dict[str, list[str]] = {
    "레킴": ["lekim", "레킴 lekim", "lekim 레킴"],
    "키코지": ["kikozy", "키코지 kikozy", "Kikozy키코지", "KIKOZY키코지"],
    "예그": ["yegg", "예그 yegg"],
    "무어니": ["MUUNI", "무어니(MUUNI)"],
    "이너프원": ["Enoughone", "이너프원 Enoughone"],
    "CDSD": ["씨디에스디"],
    "윤슬": ["Yoonseul"],
}


def collect(src: Path) -> tuple[Counter, Counter]:
    """주문 엑셀에서 공급처·판매처 값 분포를 모은다."""
    files = sorted(src.glob("*.xlsx")) if src.is_dir() else [src]
    sup, mall = Counter(), Counter()
    for path in files:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = [str(c) if c is not None else "" for c in next(rows)]
        if "공급처" not in header:
            print("[건너뜀] %s — '공급처' 컬럼이 없는 양식입니다" % path.name)
            wb.close()
            continue
        i_sup = header.index("공급처")
        i_mall = header.index("판매처") if "판매처" in header else None
        for r in rows:
            sup[str(r[i_sup] or "").strip()] += 1
            if i_mall is not None:
                mall[str(r[i_mall] or "").strip()] += 1
        wb.close()
    return sup, mall


def build(sup: Counter, mall: Counter) -> dict:
    raw_by_brand: dict[str, set] = defaultdict(set)
    for raw in sup:
        if not raw:
            continue
        key = raw
        for p in PREFIXES:
            if key.startswith(p):
                key = key[len(p):]
                break
        raw_by_brand[key.strip()].add(raw)

    brands: dict[str, dict] = {}
    for key, raws in raw_by_brand.items():
        aliases = sorted(r for r in raws if r != key)
        aliases += [a for a in EXTRA_ALIASES.get(key, []) if a not in aliases]
        names = [key] + EXTRA_ALIASES.get(key, [])
        malls = sorted(
            m for m in mall
            if m and any(n.lower() in m.lower() for n in names)
        )
        brands[key] = {
            "aliases": aliases,
            "product_keywords": [],
            "malls": malls,
            "accounts": [],
        }

    # 건수 많은 브랜드부터
    return dict(sorted(brands.items(), key=lambda kv: -sup_total(sup, kv[1], kv[0])))


def sup_total(sup: Counter, spec: dict, key: str) -> int:
    return sup.get(key, 0) + sum(sup.get(a, 0) for a in spec["aliases"])


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("src", help="주문 엑셀 폴더 또는 파일 (예: data/orders)")
    ap.add_argument("--config", default="config.json", help="갱신할 설정 파일")
    ap.add_argument("--dry-run", action="store_true", help="파일을 쓰지 않고 결과만 출력")
    args = ap.parse_args()

    src = Path(args.src).expanduser()
    if not src.is_absolute():
        src = (ROOT / src).resolve()
    cfg_path = Path(args.config).expanduser()
    if not cfg_path.is_absolute():
        cfg_path = (ROOT / cfg_path).resolve()

    if not src.exists():
        print("경로가 없습니다: %s" % src)
        return 1

    sup, mall = collect(src)
    if not sup:
        print("공급처 값을 찾지 못했습니다. '공급처' 컬럼이 있는 양식으로 다시 받아 주세요.")
        return 1

    brands = build(sup, mall)
    for k, v in brands.items():
        print("  %-12s %6d건  aliases=%d  malls=%d"
              % (k, sup_total(sup, v, k), len(v["aliases"]), len(v["malls"])))

    if args.dry_run:
        print("\n(--dry-run: 설정 파일은 건드리지 않았습니다)")
        return 0

    if not cfg_path.exists():
        print("\n설정 파일이 없습니다: %s" % cfg_path)
        print("config.example.json 을 config.json 으로 복사한 뒤 다시 실행하세요.")
        return 1

    cfg = json.loads(cfg_path.read_text(encoding="utf-8"))
    cfg["brands"] = brands
    cfg_path.write_text(json.dumps(cfg, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print("\n브랜드 %d개를 %s 에 반영했습니다." % (len(brands), cfg_path))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
