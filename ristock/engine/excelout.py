# -*- coding: utf-8 -*-
"""Ri_Stock 수집 엔진 — 엑셀 출력

원본 `기업분석표_생성.py` 의 `build_workbook` 을 그대로 옮겼습니다.
사장님이 엑셀을 계속 쓰시므로 **서식·열 순서·수식까지 원본 그대로** 유지합니다.

시트1 시총순위_분석 / 시트2 섹터별_정리 / 시트3 기업분석 / 시트4 점수기준
"""

import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .config import (ITEMS_1ST, NET5_LABELS, PICK_PER_SECTOR, TOTAL_MAX,
                     WEIGHTS, YEARS)

F_HDR = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
F_SEC = Font(name='맑은 고딕', size=11, bold=True, color='1F4E79')
F_TXT = Font(name='맑은 고딕', size=10)
FILL_HDR = PatternFill('solid', fgColor='1F4E79')
FILL_SEC = PatternFill('solid', fgColor='D9E2F3')
FILL_BAD = PatternFill('solid', fgColor='FCE4E4')
THIN = Border(*[Side(style='thin', color='BFBFBF')] * 4)


def unique_path(prefix, directory):
    """`접두어_YYYYMMDD.xlsx` — 같은 이름이 있으면 _01, _02 … (원본과 동일 규칙)"""
    os.makedirs(directory, exist_ok=True)
    date_str = pd.Timestamp.today().strftime('%Y%m%d')
    path = os.path.join(directory, f'{prefix}_{date_str}.xlsx')
    k = 0
    while os.path.exists(path):
        k += 1
        path = os.path.join(directory, f'{prefix}_{date_str}_{k:02d}.xlsx')
    return path


def write_header(ws, row, headers):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = F_HDR
        c.fill = FILL_HDR
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = THIN


def write_row(ws, row, values, num_fmts=None):
    for j, v in enumerate(values, 1):
        c = ws.cell(row=row, column=j, value=v)
        c.font = F_TXT
        c.border = THIN
        if num_fmts and j in num_fmts and isinstance(v, (int, float)):
            c.number_format = num_fmts[j]


RANK_COLS = ['시총순위', '티커', '기업명', '한글명', '섹터', '산업', '국가', '거래소', '시가총액(조원)',
             'PER(배)', 'PBR(배)', 'ROE(%)', 'PEG', 'EV/EBITDA', 'PSR', '배당수익률(%)',
             '부채비율(%)', '평가점수(1차·100환산)', '3년 평균 주가 수익률', '5년 평균 주가 수익률',
             '10년 평균 주가 수익률'] + [str(y) for y in YEARS]


def rank_row_values(r):
    return [r['시총순위'], r['티커'], r['기업명'], r.get('한글명', ''), r['섹터'],
            r['산업'], r['국가'], r['거래소'], round(r['시총_조원'], 1),
            r['PER'], r['PBR'], r['ROE'], r['PEG'], r['EV/EBITDA'], r['PSR'],
            r['배당수익률'], r['부채비율'], r['평가점수(1차)'],
            r['3년수익률'], r['5년수익률'], r['10년수익률']] + [r[str(y)] for y in YEARS]


def build_workbook(df, market_label, out_path):
    wb = Workbook()
    fmt_1 = '#,##0.0'
    fmt_3 = '0.000'
    num_fmts = {9: '#,##0.0', 10: fmt_1, 11: fmt_1, 12: fmt_1, 13: fmt_1, 14: fmt_1,
                15: fmt_1, 16: fmt_1, 17: fmt_1, 18: fmt_1, 19: fmt_3, 20: fmt_3, 21: fmt_3}
    for j in range(22, 22 + len(YEARS)):
        num_fmts[j] = '#,##0.00'

    # ---- 시트1: 시총순위_분석 ----
    ws = wb.active
    ws.title = '시총순위_분석'
    write_header(ws, 1, RANK_COLS)
    for i, (_, r) in enumerate(df.sort_values('시총순위').iterrows(), 2):
        write_row(ws, i, rank_row_values(r), num_fmts)
    ws.freeze_panes = 'E2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(RANK_COLS))}{len(df) + 1}'
    widths = [8, 10, 26, 16, 13, 24, 10, 10, 12] + [8] * 9 + [11, 11, 11] + [9] * len(YEARS)
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # ---- 시트2: 섹터별_정리 ----
    ws2 = wb.create_sheet('섹터별_정리')
    row = 1
    sec_order = (df.groupby('섹터')['시총_조원'].sum().sort_values(ascending=False).index)
    for sec in sec_order:
        g = df[df['섹터'] == sec].sort_values('시총순위')
        c = ws2.cell(row=row, column=1,
                     value=f'■ {sec}  ({len(g)}개 · 시총 합계 {g["시총_조원"].sum():,.0f}조원)')
        c.font = F_SEC
        for j in range(1, len(RANK_COLS) + 1):
            ws2.cell(row=row, column=j).fill = FILL_SEC
        row += 1
        write_header(ws2, row, RANK_COLS)
        row += 1
        for _, r in g.iterrows():
            write_row(ws2, row, rank_row_values(r), num_fmts)
            row += 1
        row += 1
    for j, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(j)].width = w

    # ---- 시트3: 기업분석 (항목 10점 만점 표시 + 가중치 자동 총점) ----
    ws3 = wb.create_sheet('기업분석')
    score_items = ITEMS_1ST + ['뉴스', '수급']          # 13개 점수 열
    extra_cols = ['미네르비니', '모멘텀12M(%)', '거래량비(%)', '순익성장률(%)',
                  'ROE(%)', 'PER(배)']
    cols3 = (['섹터내순위', '시총순위', '티커', '기업명', '한글명', '총점(100)']
             + [f'점수_{it}' for it in score_items]
             + extra_cols
             + ['순익연도'] + NET5_LABELS
             + ['모멘텀근거', '차트근거', '매출근거', '순익근거', '재무근거', '52주근거',
                '변동근거', '밸류근거', '배당근거', '수급근거', '미네르비니근거',
                '호재요약', '악재사유'])
    TOTAL_COL = 6                                    # 총점(100) 열(F)
    SCORE_C0 = 7                                     # 첫 점수 열(G)
    SCORE_C1 = SCORE_C0 + len(score_items) - 1       # 마지막 점수 열(R)
    col_f = get_column_letter(SCORE_C0)
    col_l = get_column_letter(SCORE_C1)
    NET_C0 = SCORE_C1 + 2 + len(extra_cols)          # 순익 첫 열
    nf3 = {TOTAL_COL: fmt_1}
    for j in range(SCORE_C0, SCORE_C1 + 1):
        nf3[j] = '0.0'
    for j in range(SCORE_C1 + 2, SCORE_C1 + 2 + len(extra_cols) - 1):
        nf3[j] = '0.0'                               # 모멘텀12M~PER
    for j in range(NET_C0, NET_C0 + len(NET5_LABELS)):
        nf3[j] = '#,##0'

    row = 1
    evaluated = df[df['총점(100)'].notna()]
    for sec in sec_order:
        g = evaluated[evaluated['섹터'] == sec]
        if g.empty:
            continue
        g = g.sort_values('총점(100)', ascending=False)
        c = ws3.cell(row=row, column=1,
                     value=f'■ {sec}  (평가 {len(g)}종목 · 항목별 10점 만점, 총점은 가중치 자동합산 후 100점 환산)')
        c.font = F_SEC
        for j in range(1, len(cols3) + 1):
            ws3.cell(row=row, column=j).fill = FILL_SEC
        row += 1
        write_header(ws3, row, cols3)
        row += 1
        # 가중치 행 (총점 수식이 이 행을 참조)
        wrow = row
        ws3.cell(row=wrow, column=TOTAL_COL - 1, value='가중치 →').font = F_TXT
        ws3.cell(row=wrow, column=TOTAL_COL, value=f'만점 {TOTAL_MAX}→100').font = F_TXT
        for j, it in enumerate(score_items, SCORE_C0):
            cell = ws3.cell(row=wrow, column=j, value=WEIGHTS[it])
            cell.font = Font(name='맑은 고딕', size=9, italic=True, color='808080')
            cell.border = THIN
        row += 1
        for rank, (_, r) in enumerate(g.iterrows(), 1):
            vals = ([rank, r['시총순위'], r['티커'], r['기업명'], r.get('한글명', ''), None]
                    + [r[f'점수_{it}'] for it in score_items]
                    + [r.get(c) for c in extra_cols]
                    + [r['순익연도']] + [r[lab] for lab in NET5_LABELS]
                    + [r['모멘텀근거'], r['차트근거'], r['매출근거'], r['순익근거'],
                       r['재무근거'], r['52주근거'], r['변동근거'], r['밸류근거'],
                       r['배당근거'], r['수급근거'], r.get('미네르비니근거', ''),
                       r['호재요약'], r['악재사유']])
            write_row(ws3, row, vals, nf3)
            # 총점 = SUMPRODUCT(항목10점점수, 가중치)/만점×100 (점수를 고치면 자동 재계산)
            tc = ws3.cell(row=row, column=TOTAL_COL,
                          value=f'=ROUND(SUMPRODUCT({col_f}{row}:{col_l}{row},'
                                f'{col_f}${wrow}:{col_l}${wrow})*10/{TOTAL_MAX},1)')
            tc.font = Font(name='맑은 고딕', size=10, bold=True)
            tc.number_format = fmt_1
            tc.border = THIN
            if r['악재탐지']:
                for j in range(1, len(cols3) + 1):
                    ws3.cell(row=row, column=j).fill = FILL_BAD
            row += 1
        row += 1
    w3 = ([9, 8, 10, 24, 16, 11] + [8] * len(score_items) + [10] * len(extra_cols)
          + [12] + [12] * len(NET5_LABELS)
          + [22, 30, 13, 13, 16, 13, 13, 16, 14, 20, 28, 40, 40])
    for j, w in enumerate(w3, 1):
        ws3.column_dimensions[get_column_letter(j)].width = w
    ws3.freeze_panes = 'G1'

    # ---- 시트4: 점수기준 ----
    ws4 = wb.create_sheet('점수기준')
    criteria = [
        ['항목', '가중치', '기준 및 근거 (기업분석 시트의 점수는 모두 10점 만점 표시)'],
        ['⭐ 12개월 모멘텀', 12, '12-1 모멘텀 60%↑=10, 30~60%=7.5, 0~30%=4.2, 음수=0 [한국시장 백테스트: 60%↑ 구간 20일 평균 +5.3% vs 음수 +2.9%]'],
        ['① 눌림목·차트', 12, '최근 10일 고점 대비 -1.5% 이상 눌림 시 — 1차 눌림(MA20 ±3% 근접)+정배열=10 / 2차 눌림(MA120·240·448 근접)=6.7 / 지지선 미근접=4.2 / 미눌림: 정배열 3.3, 역배열 0.8'],
        ['⭐ ROE(수익성)', 8, 'ROE 15%↑=10, 10%↑=7.5, 5%↑=5 [세계 퀄리티 팩터, Novy-Marx 2013]'],
        ['② 매출 성장', 8, '최근 3년 중 증가 구간 2개=10, 1개=6.3'],
        ['③ 순이익 성장', 8, '최근 3년 중 증가 구간 2개=10, 1개=6.3'],
        ['④ 시가총액', 5, '10조↑=10, 1조↑=8, 5천억↑=6, 1200억↑=4 (원화 환산 기준) [사이즈 팩터]'],
        ['⑤ 재무(부채+흑자)', 12, '부채비율 50%↓~250%↑ 구간점수 + 최근 순이익 흑자 가점'],
        ['⭐ 밸류(저평가)', 8, '저PER(10↓=만점~40↓) + 저PBR(1↓=만점~8↓) 각 절반 [Fama-French 1992 가치 팩터]'],
        ['배당수익률', 4, '4%↑=10, 2%↑=7.5, 1%↑=5, 0 초과=2.5 [배당 프리미엄 팩터]'],
        ['⑦ 호재 뉴스', 12, '뉴스 제목 규칙분석 호재점수 0~10점 그대로 표시 / 치명적 악재 종목은 행 표시'],
        ['⑧ 수급', 13, '한국: 최근 5일 (기관+외인 순매수)÷거래량 비율 / 미국: 데이터 없음 → 중립 3.1'],
        ['⑨ 52주 위치', 5, '고점대비 75%↑=10, 40%미만=4, 55~75%=0 [백테스트: 75%↑ 최고, 55~75% 최악]'],
        ['⭐ 저변동성', 5, '20일 변동성 2%↓=10, 3%↓=8, 4%↓=4 [백테스트: 저변동 승률 55.3% vs 고변동 49.1%]'],
        ['', '', ''],
        ['총점 계산', '', f'총점(100) = Σ(항목 10점점수 × 가중치) ÷ 10 ÷ {TOTAL_MAX} × 100 — 기업분석 시트에 엑셀 수식으로 들어있어 점수를 수정하면 자동 재계산됩니다.'],
        ['순익 5년', '', '순익최근(억원) 등 5칸은 최근 회계연도 순이익(원화 억원 환산). 데이터 소스(야후 파이낸스)가 통상 4개 회계연도까지 제공해 가장 오래된 칸은 비어있을 수 있습니다.'],
        ['참고', '', f'{market_label} 시총 상위 {len(df)}종목 대상. 뉴스/수급은 섹터별 1차점수 상위 {PICK_PER_SECTOR}종목만 분석.'],
        ['', '', '본 자료는 스크리닝 참고용이며 투자 권유가 아닙니다. 투자 판단과 책임은 본인에게 있습니다.'],
    ]
    for i, rowvals in enumerate(criteria, 1):
        for j, v in enumerate(rowvals, 1):
            c = ws4.cell(row=i, column=j, value=v)
            c.font = F_HDR if i == 1 else F_TXT
            if i == 1:
                c.fill = FILL_HDR
    ws4.column_dimensions['A'].width = 20
    ws4.column_dimensions['B'].width = 8
    ws4.column_dimensions['C'].width = 110

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)
    print(f'  저장 완료: {out_path}')
    return out_path
