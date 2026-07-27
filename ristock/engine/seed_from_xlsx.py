# -*- coding: utf-8 -*-
"""기존 기업분석표 xlsx → 설계서 2.2 규격 JSON 변환기 (초기 시드 생성용)

왜 필요한가
-----------
앱이 **첫날부터 진짜 데이터로** 열려야 합니다. 새 수집 엔진이 처음 돌기 전까지는
사장님이 이미 만들어 둔 `한국/미국_기업_분석표_자동_YYYYMMDD.xlsx` 가 유일한 실데이터입니다.
이 스크립트는 그 xlsx 를 `ristock/data/stocks_KR.json` · `stocks_US.json` 으로 바꿉니다.

절대 원칙
---------
* **계산 로직을 새로 만들지 않습니다.** 총점은 시트에 적힌 값을 쓰지 않고,
  `주식_대시보드/app.py` 의 `parse_analysis_sheet()` 와 **글자 그대로 같은 식**으로 재계산합니다.
  (시트의 총점 칸은 엑셀 수식이라 `data_only=True` 로 읽으면 비어 있습니다.)
* `지표.고점대비` / `지표.10일고점비` 는 옛 xlsx 에 숫자 컬럼이 없어
  `52주근거` · `차트근거` 문장에서 뽑습니다. 정규식은 app.py 의
  `high_position()` / `dd10()` 과 **완전히 동일**합니다.
* 종목 배열 순서는 xlsx `기업분석` 시트 순서(섹터별 → 총점 내림차순)를 그대로 유지합니다.
  동점 종목의 순서가 어긋나면 기존 전략 결과(정답지)를 재현할 수 없습니다.

사용법
------
    # xlsx → 종목 JSON (통계까지 함께 출력)
    python seed_from_xlsx.py build --xlsx 한국_기업_분석표_자동_20260716.xlsx \
        --market kr --date 20260716 --generated "2026-07-16 13:48" \
        --out ../data/stocks_KR.json --stats

    # 만들어진 JSON 되짚어 검증 (총점 일치율 100% 여야 정상)
    python seed_from_xlsx.py verify --json ../data/stocks_KR.json \
        --xlsx 한국_기업_분석표_자동_20260716.xlsx

    # 통계만 다시 보기
    python seed_from_xlsx.py stats --json ../data/stocks_KR.json

    # manifest.json 생성
    python seed_from_xlsx.py manifest --kr ../data/stocks_KR.json \
        --us ../data/stocks_US.json --out ../data/manifest.json

    # strategies.json 생성 (원본 대시보드 strategies.json + 섹터매핑 + 필터설명)
    python seed_from_xlsx.py strategies --src 주식_대시보드/strategies.json \
        --out ../data/strategies.json
"""

import argparse
import json
import math
import os
import re
import sys
import time

from openpyxl import load_workbook

# =========================
# 시트 컬럼 ↔ JSON 필드 (설계서 2.2)
# =========================
MARKET_NAME = {'kr': '한국', 'us': '미국'}

# 13개 점수 항목 — 엑셀 `점수_XXX` 열 순서와 동일해야 합니다.
SCORE_ITEMS = ['모멘텀', '차트', 'ROE', '매출성장', '순익성장', '시총', '재무',
               '52주', '저변동', '밸류', '배당', '뉴스', '수급']

# 근거 JSON 키 → 엑셀 컬럼명
REASON_COLS = {
    '모멘텀': '모멘텀근거', '차트': '차트근거', '매출': '매출근거', '순익': '순익근거',
    '재무': '재무근거', '52주': '52주근거', '변동': '변동근거', '밸류': '밸류근거',
    '배당': '배당근거', '수급': '수급근거', '미네르비니': '미네르비니근거',
}

# 지표 JSON 키 → `기업분석` 시트 컬럼명 (2026-07-16 이후 서식에만 있는 열도 포함)
METRIC_FROM_ANALYSIS = {
    '모멘텀12M': '모멘텀12M(%)',
    '거래량비': '거래량비(%)',
    '순익성장률': '순익성장률(%)',
}
# 지표 JSON 키 → `시총순위_분석` 시트 컬럼명 (300종목 전체가 가진 값)
METRIC_FROM_UNIVERSE = {
    'PER': 'PER(배)', 'PBR': 'PBR(배)', 'ROE': 'ROE(%)',
    '배당수익률': '배당수익률(%)', '부채비율': '부채비율(%)',
}
# 지표 키 전체 순서 (없으면 null 로라도 반드시 채웁니다 — 화면·전략이 키 존재를 가정)
METRIC_KEYS = ['PER', 'PBR', 'ROE', '배당수익률', '부채비율',
               '모멘텀12M', '거래량비', '순익성장률', '고점대비', '10일고점비']

NET5_LABELS = ['순익-4(억원)', '순익-3(억원)', '순익-2(억원)', '순익-1(억원)', '순익최근(억원)']
RETURN_COLS = {'3년': '3년 평균 주가 수익률', '5년': '5년 평균 주가 수익률',
               '10년': '10년 평균 주가 수익률'}
YEARS = [str(y) for y in range(2015, 2026)]

# 뉴스 섹터 → 분석표 섹터 (원본 app.py 의 SECTOR_MAP_KR / SECTOR_MAP_US 를 그대로 옮김)
SECTOR_MAP_KR = {
    'AI·반도체': ['반도체', 'IT·인터넷·게임', '전자·부품'],
    '로봇·자동화': ['조선·기계·방산', '전자·부품'],
    '방산·우주': ['조선·기계·방산'],
    '원전·전력인프라': ['전기·전력'],
    '2차전지·전기차': ['2차전지·화학', '자동차'],
    '바이오·제약': ['바이오·제약'],
    '조선·해운': ['조선·기계·방산', '운송·물류'],
    '금융·은행': ['금융·지주'],
    '에너지·정유': ['에너지·정유'],
    '소비·유통': ['유통·소비재'],
    '엔터·게임·콘텐츠': ['엔터·미디어', 'IT·인터넷·게임'],
    '수소·신재생에너지': ['전기·전력', '2차전지·화학'],
    '철강·소재': ['철강·소재'],
}
SECTOR_MAP_US = {
    'AI·반도체': ['IT·기술'],
    '로봇·자동화': ['산업재', 'IT·기술'],
    '방산·우주': ['산업재'],
    '원전·전력인프라': ['유틸리티', '산업재'],
    '2차전지·전기차': ['경기소비재', 'IT·기술'],
    '바이오·제약': ['헬스케어'],
    '조선·해운': ['산업재'],
    '금융·은행': ['금융'],
    '에너지·정유': ['에너지'],
    '소비·유통': ['필수소비재', '경기소비재'],
    '엔터·게임·콘텐츠': ['커뮤니케이션'],
    '수소·신재생에너지': ['유틸리티', '에너지'],
    '철강·소재': ['소재'],
}

# 설계서 3.1 필터 표를 사장님이 읽을 한 줄 설명으로 옮긴 것
FILTER_DESC = {
    '악재제외': '치명적 악재 뉴스가 잡힌 종목은 제외합니다',
    '흑자만': '최근 순이익이 흑자인 종목만 남깁니다 (재무 근거에 "흑자"가 있어야 통과)',
    '눌림만': '최근 10일 고점에서 밀려 눌림 상태인 종목만 남깁니다 (미눌림은 제외)',
    '전고점돌파만': '52주 고점의 85% 이상이면서 최근 10일 고점을 뚫고 오르는 중인 종목만 남깁니다',
    '미네르비니만': '미네르비니 트렌드 템플릿(SEPA)을 통과한 종목만 남깁니다',
    '거래량150이상': '거래량이 평소의 150% 이상으로 늘어난 종목만 남깁니다',
    '순익성장25이상': '연간 순이익이 25% 이상 성장한 종목만 남깁니다',
    '신고가근접만': '52주 고점의 75% 이상까지 올라온 종목만 남깁니다',
}

MAX_JSON_BYTES = 400 * 1024   # 폰에서 받는 파일 — 이 크기를 넘으면 연도별종가를 줄입니다


# =========================
# 원본 app.py 와 동일한 파싱 (한 줄도 바꾸지 않음)
# =========================
def parse_analysis_sheet(path):
    """`기업분석` 시트의 섹터 블록 → {섹터: [종목dict...]} (총점은 가중치로 재계산)

    app.py 의 같은 이름 함수를 그대로 옮긴 것입니다. 조건문 순서까지 동일합니다.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb['기업분석']
    sectors = {}
    cur_sector = None
    header = []
    weights = []
    score_cols = []

    for row in ws.iter_rows(values_only=True):
        a = row[0]
        if isinstance(a, str) and a.startswith('■'):
            cur_sector = a[1:].split('(')[0].strip()
            header, weights = [], []
            continue
        if a == '섹터내순위':
            header = list(row)
            score_cols = [i for i, h in enumerate(header)
                          if isinstance(h, str) and h.startswith('점수_')]
            continue
        if header and not weights and a is None and score_cols and \
                any(isinstance(row[i], (int, float)) for i in score_cols):
            weights = [row[i] for i in score_cols]
            continue
        if cur_sector and header and isinstance(a, (int, float)):
            rec = {h: row[i] for i, h in enumerate(header) if h}
            if weights and score_cols:
                sp = sum((rec.get(header[i]) or 0) * w
                         for i, w in zip(score_cols, weights) if w)
                wsum = sum(w for w in weights if w)
                rec['총점(100)'] = round(sp / 10 / wsum * 100, 1) if wsum else None
            sectors.setdefault(cur_sector, []).append(rec)
    return sectors


def _parse_pct(text, pattern):
    m = re.search(pattern, str(text or ''))
    return float(m.group(1)) if m else None


def high_position(rec):
    """52주 고점대비 위치(%) — 52주근거 '고점대비 xx%' 에서 추출 (app.py 와 동일)"""
    return _parse_pct(rec.get('52주근거'), r'고점대비 (-?\d+(?:\.\d+)?)%')


def dd10(rec):
    """최근 10일 고점 대비 등락(%) — 차트근거에서 추출 (app.py 와 동일, 양수 = 신고가 경신 중)"""
    return _parse_pct(rec.get('차트근거'), r'10일고점比 (-?\d+(?:\.\d+)?)%')


def read_weights(path):
    """`기업분석` 시트의 가중치 행 → {'모멘텀':12, ...}

    가중치는 파일마다 하드코딩하지 않고 **시트에서 직접 읽습니다.**
    섹터 블록마다 반복되는데, 하나라도 다르면 총점 재계산이 틀어지므로 즉시 실패시킵니다.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb['기업분석']
    header, weights, score_cols = [], [], []
    found = []
    for row in ws.iter_rows(values_only=True):
        a = row[0]
        if isinstance(a, str) and a.startswith('■'):
            header, weights = [], []
            continue
        if a == '섹터내순위':
            header = list(row)
            score_cols = [i for i, h in enumerate(header)
                          if isinstance(h, str) and h.startswith('점수_')]
            continue
        if header and not weights and a is None and score_cols and \
                any(isinstance(row[i], (int, float)) for i in score_cols):
            weights = [row[i] for i in score_cols]
            found.append({header[i][3:]: w for i, w in zip(score_cols, weights)})
    if not found:
        raise SystemExit(f'가중치 행을 찾지 못했습니다: {path}')
    if any(w != found[0] for w in found):
        raise SystemExit(f'섹터 블록마다 가중치가 다릅니다 — 원본 확인 필요: {path}')
    if list(found[0].keys()) != SCORE_ITEMS:
        raise SystemExit(f'점수 항목 구성이 예상과 다릅니다: {list(found[0].keys())}')
    return found[0]


def parse_universe_sheet(path):
    """`시총순위_분석` 시트 → {티커: 기본정보dict} (시총순위 오름차순, 입력 순서 유지)"""
    wb = load_workbook(path, data_only=True)
    ws = wb['시총순위_분석']
    rows = ws.iter_rows(values_only=True)
    header = [h for h in next(rows)]
    idx = {h: i for i, h in enumerate(header) if h}
    out = {}
    for row in rows:
        ticker = row[idx['티커']]
        if ticker is None:
            continue
        out[str(ticker)] = {h: row[i] for h, i in idx.items()}
    return out


# =========================
# 값 다듬기
# =========================
def _num(v):
    """엑셀 값 → JSON 숫자. 빈칸·NaN·문자열 쓰레기는 모두 None."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        return v
    s = str(v).strip().replace(',', '')
    if not s or s.lower() in ('nan', 'inf', '-inf', '-', 'none'):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _text(v):
    """엑셀 값 → JSON 문자열. 빈칸은 빈 문자열.

    **앞뒤 공백을 다듬지 않습니다.** 원본 `호재요약` 에는 앞에 공백이 붙은 칸이 있는데,
    기존 대시보드는 이 문장을 `[:120]` 으로 잘라 화면에 씁니다.
    공백을 지우면 자르는 위치가 한 칸씩 밀려 정답지와 글자가 어긋납니다.
    """
    if v is None:
        return ''
    return str(v)


# =========================
# 스냅샷(= stocks_XX.json) 조립
# =========================
def build_snapshot(xlsx_path, market, base_date, generated_at=None, closes='auto'):
    """xlsx 한 개 → 설계서 2.2 규격 dict

    market      : '한국' | '미국'
    base_date   : '20260716'
    generated_at: '2026-07-16 13:48' (생략하면 파일 수정시각)
    closes      : 'all'(전 종목 연도별종가) | 'evaluated'(평가 종목만) | 'auto'(400KB 넘으면 자동 축소)
    """
    weights = read_weights(xlsx_path)
    sectors = parse_analysis_sheet(xlsx_path)
    universe = parse_universe_sheet(xlsx_path)

    if generated_at is None:
        generated_at = time.strftime('%Y-%m-%d %H:%M',
                                     time.localtime(os.path.getmtime(xlsx_path)))

    stocks = []
    evaluated = set()

    # (1) 평가 종목 — `기업분석` 시트 순서 그대로 (섹터별 → 총점 내림차순)
    for sector, rows in sectors.items():
        for rec in rows:
            ticker = _text(rec.get('티커'))
            uni = universe.get(ticker, {})
            stocks.append(_evaluated_record(ticker, sector, rec, uni))
            evaluated.add(ticker)

    # (2) 평가 안 된 나머지 — 시총순위 순으로 뒤에 붙임
    for ticker, uni in universe.items():
        if ticker in evaluated:
            continue
        stocks.append(_plain_record(ticker, uni))

    snap = {
        '시장': market,
        '기준일': base_date,
        '생성시각': generated_at,
        '원본': os.path.basename(xlsx_path),
        '가중치': weights,
        '종목': stocks,
    }

    if closes == 'evaluated':
        _trim_closes(snap)
    elif closes == 'auto' and _json_bytes(snap) > MAX_JSON_BYTES:
        _trim_closes(snap)
    return snap


def _evaluated_record(ticker, sector, rec, uni):
    """평가 종목 한 줄 — 점수·근거·지표·미네르비니·호재·악재·순익 5칸까지 채웁니다."""
    지표 = {k: None for k in METRIC_KEYS}
    for key, col in METRIC_FROM_UNIVERSE.items():
        지표[key] = _num(uni.get(col))
    for key, col in METRIC_FROM_ANALYSIS.items():
        지표[key] = _num(rec.get(col))
    # 옛 서식에는 숫자 컬럼이 없어 근거 문장에서 뽑습니다 (app.py 와 같은 정규식)
    지표['고점대비'] = high_position(rec)
    지표['10일고점비'] = dd10(rec)

    out = {
        '티커': ticker,
        '기업명': _text(rec.get('기업명') or uni.get('기업명')),
        '한글명': _text(rec.get('한글명') or uni.get('한글명')),
        '섹터': sector,
        '산업': _text(uni.get('산업')),
        '거래소': _text(uni.get('거래소')),
        '시총순위': _num(rec.get('시총순위') if rec.get('시총순위') is not None
                      else uni.get('시총순위')),
        '시총조원': _num(uni.get('시가총액(조원)')),
        '평가': True,
        '총점': rec.get('총점(100)'),
        '점수': {k: _num(rec.get(f'점수_{k}')) for k in SCORE_ITEMS},
        '근거': {k: _text(rec.get(col)) for k, col in REASON_COLS.items()},
        '지표': 지표,
        '미네르비니': _text(rec.get('미네르비니')),
        '호재': _text(rec.get('호재요약')),
        '악재': _text(rec.get('악재사유')),
        '순익': {'연도': _text(rec.get('순익연도')),
                '값': [_num(rec.get(c)) for c in NET5_LABELS]},
        '연도별종가': _closes(uni),
        '수익률': {k: _num(uni.get(col)) for k, col in RETURN_COLS.items()},
    }
    return out


def _plain_record(ticker, uni):
    """평가 안 된 종목 — 뉴스·수급을 돌리지 않았으므로 점수·근거·총점이 없습니다(설계서 2.2).

    ⚠ **키 구성은 `engine/emit.stock_record()` 가 정본입니다.**
    같은 `stocks_XX.json` 을 만드는 생산자가 둘(시드 변환기·수집 엔진)인데 필드가 서로
    달라지면, 시드로 열던 앱이 첫 수집 뒤 갑자기 필드가 생기는 불안정한 계약이 됩니다.
    그래서 `미네르비니 · 호재 · 악재 · 순익` 을 미평가 종목에도 **키만이라도** 채웁니다.

    옛 xlsx 의 `시총순위_분석` 시트에는 이 네 값이 아예 없습니다
    (뉴스·수급을 돌린 종목만 `기업분석` 시트에 실렸기 때문입니다).
    읽을 수 없는 값은 **키는 두고 빈 값**으로 둡니다 — 없는 값을 지어내지 않습니다.
    `지표.모멘텀12M · 거래량비 · 순익성장률` 도 같은 이유로 null 입니다.
    """
    지표 = {k: None for k in METRIC_KEYS}
    for key, col in METRIC_FROM_UNIVERSE.items():
        지표[key] = _num(uni.get(col))
    return {
        '티커': ticker,
        '기업명': _text(uni.get('기업명')),
        '한글명': _text(uni.get('한글명')),
        '섹터': _text(uni.get('섹터')),
        '산업': _text(uni.get('산업')),
        '거래소': _text(uni.get('거래소')),
        '시총순위': _num(uni.get('시총순위')),
        '시총조원': _num(uni.get('시가총액(조원)')),
        '평가': False,
        '지표': 지표,
        '미네르비니': _text(uni.get('미네르비니')),
        '호재': _text(uni.get('호재요약')),
        '악재': _text(uni.get('악재사유')),
        '순익': {'연도': _text(uni.get('순익연도')),
                '값': [_num(uni.get(c)) for c in NET5_LABELS]},
        '연도별종가': _closes(uni),
        '수익률': {k: _num(uni.get(col)) for k, col in RETURN_COLS.items()},
    }


def _closes(uni):
    """연도별 종가 — 값이 있는 해만 담습니다 (빈 해까지 넣으면 파일만 커집니다)."""
    out = {}
    for y in YEARS:
        v = _num(uni.get(y) if y in uni else uni.get(int(y)))
        if v is not None:
            out[y] = v
    return out


def _trim_closes(snap):
    """평가 안 된 종목의 연도별종가를 비웁니다 (파일 크기 줄이기 — 설계 지시)."""
    for s in snap['종목']:
        if not s.get('평가'):
            s['연도별종가'] = {}


def _json_bytes(obj):
    return len(dump_json(obj).encode('utf-8'))


def dump_json(obj):
    """앱이 내려받는 파일이라 공백 없이 저장합니다 (한글은 그대로)."""
    return json.dumps(obj, ensure_ascii=False, separators=(',', ':'))


def write_json(path, obj):
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        f.write(dump_json(obj))
    return os.path.getsize(path)


# =========================
# 통계 · 역검증
# =========================
def stats_text(snap, path=None):
    stocks = snap['종목']
    ev = [s for s in stocks if s.get('평가')]
    totals = [s['총점'] for s in ev if s.get('총점') is not None]
    sectors = []
    for s in ev:
        if s['섹터'] not in sectors:
            sectors.append(s['섹터'])
    hp_null = sum(1 for s in ev if s['지표'].get('고점대비') is None)
    dd_null = sum(1 for s in ev if s['지표'].get('10일고점비') is None)
    size = os.path.getsize(path) if path and os.path.exists(path) else _json_bytes(snap)
    lines = [
        f"[{snap['시장']} {snap['기준일']}] 원본 {snap['원본']}  생성시각 {snap['생성시각']}",
        f"  종목수 {len(stocks)} · 평가종목수 {len(ev)} · 미평가 {len(stocks) - len(ev)}",
        f"  섹터 {len(sectors)}개: {', '.join(sectors)}",
        f"  총점 최소 {min(totals):.1f} / 최대 {max(totals):.1f} / 평균 "
        f"{sum(totals) / len(totals):.2f}" if totals else '  총점 없음',
        f"  지표.고점대비 null {hp_null}개 · 지표.10일고점비 null {dd_null}개 (평가종목 기준)",
        f"  파일크기 {size:,} 바이트 ({size / 1024:.1f} KB) — 한도 {MAX_JSON_BYTES // 1024} KB",
    ]
    return '\n'.join(lines)


def verify_totals(snap, xlsx_path):
    """역검증 — JSON 으로 app.py 의 rec dict 를 되짚어 만들고 총점을 다시 계산해 비교합니다.

    JSON 만 가지고 원본 총점을 소수 1자리까지 재현할 수 있어야 합니다(=100%).
    """
    weights = snap['가중치']
    wsum = sum(w for w in weights.values() if w)
    ok = bad = 0
    diffs = []
    for s in snap['종목']:
        if not s.get('평가'):
            continue
        sp = sum((s['점수'].get(k) or 0) * weights[k] for k in SCORE_ITEMS if weights[k])
        again = round(sp / 10 / wsum * 100, 1) if wsum else None
        if again == s['총점']:
            ok += 1
        else:
            bad += 1
            diffs.append((s['티커'], s['총점'], again))

    # 원본 xlsx 와도 대조 (JSON 이 시트에서 벗어나지 않았는지)
    src_ok = src_bad = 0
    src_diffs = []
    if xlsx_path:
        sectors = parse_analysis_sheet(xlsx_path)
        src = {}
        for rows in sectors.values():
            for rec in rows:
                src[str(rec.get('티커'))] = rec
        by_ticker = {s['티커']: s for s in snap['종목'] if s.get('평가')}
        for t, rec in src.items():
            j = by_ticker.get(t)
            if j is not None and j['총점'] == rec['총점(100)']:
                src_ok += 1
            else:
                src_bad += 1
                src_diffs.append((t, rec['총점(100)'], j['총점'] if j else None))
    return {'재계산일치': ok, '재계산불일치': bad, '불일치목록': diffs[:10],
            '원본일치': src_ok, '원본불일치': src_bad, '원본불일치목록': src_diffs[:10]}


# =========================
# manifest / strategies
# =========================
def build_manifest(kr_path, us_path, runner='pc', memo=''):
    """manifest.json — 시장별 `기준일` 을 각각 담고, 최상단은 그중 **가장 오래된** 값.

    앱의 신선도 판정은 최상단 `기준일` 하나만 봅니다. 한쪽 시장만 낡았을 때
    새 쪽 날짜를 올려 두면 낡은 데이터가 "오늘 것"으로 보입니다
    (`engine/emit.대표기준일()` 과 같은 규칙입니다).
    """
    시장 = {}
    생성시각 = ''
    기준일들 = []
    for market, path in [('한국', kr_path), ('미국', us_path)]:
        if not path:
            continue
        with open(path, encoding='utf-8') as f:
            snap = json.load(f)
        ev = sum(1 for s in snap['종목'] if s.get('평가'))
        시장[market] = {
            '파일': os.path.basename(path),
            '종목수': len(snap['종목']),
            '평가종목수': ev,
            '기준일': snap['기준일'],
            '원본': snap['원본'],
            '수집시각': snap['생성시각'],
        }
        기준일들.append(snap['기준일'])
        생성시각 = max(생성시각, snap['생성시각'])
    기준일 = min(기준일들) if 기준일들 else ''
    return {
        '데이터버전': 1,
        '생성시각': 생성시각,
        '기준일': 기준일,
        '시장': 시장,
        '실행': {'주체': runner, '성공': True, '메모': memo},
    }


def build_strategies(src_path):
    with open(src_path, encoding='utf-8') as f:
        strategies = json.load(f)
    return {
        '전략': strategies,
        '섹터매핑': {'한국': SECTOR_MAP_KR, '미국': SECTOR_MAP_US},
        '필터설명': FILTER_DESC,
    }


# =========================
# CLI
# =========================
def main():
    ap = argparse.ArgumentParser(description='기업분석표 xlsx → 설계서 2.2 JSON 변환기')
    sub = ap.add_subparsers(dest='cmd', required=True)

    b = sub.add_parser('build', help='xlsx → stocks_XX.json')
    b.add_argument('--xlsx', required=True)
    b.add_argument('--market', choices=['kr', 'us'], required=True)
    b.add_argument('--date', required=True, help='기준일 YYYYMMDD')
    b.add_argument('--generated', default=None, help='생성시각 "YYYY-MM-DD HH:MM" (생략 시 파일 수정시각)')
    b.add_argument('--out', required=True)
    b.add_argument('--closes', choices=['all', 'evaluated', 'auto'], default='auto',
                   help='연도별종가 범위 (auto: 400KB 초과 시 평가 종목만)')
    b.add_argument('--stats', action='store_true', help='변환 후 통계 출력')

    v = sub.add_parser('verify', help='JSON ↔ xlsx 총점 역검증')
    v.add_argument('--json', required=True)
    v.add_argument('--xlsx', default=None)

    s = sub.add_parser('stats', help='JSON 통계만 출력')
    s.add_argument('--json', required=True)

    m = sub.add_parser('manifest', help='manifest.json 생성')
    m.add_argument('--kr', required=True)
    m.add_argument('--us', required=True)
    m.add_argument('--out', required=True)
    m.add_argument('--runner', default='pc', choices=['pc', 'github-actions'])
    m.add_argument('--memo', default='')

    t = sub.add_parser('strategies', help='strategies.json 생성')
    t.add_argument('--src', required=True, help='원본 주식_대시보드/strategies.json')
    t.add_argument('--out', required=True)

    a = ap.parse_args()

    if a.cmd == 'build':
        snap = build_snapshot(a.xlsx, MARKET_NAME[a.market], a.date, a.generated, a.closes)
        size = write_json(a.out, snap)
        print(f'저장: {a.out} ({size:,} 바이트)')
        if a.stats:
            print(stats_text(snap, a.out))
        return

    if a.cmd == 'stats':
        with open(a.json, encoding='utf-8') as f:
            snap = json.load(f)
        print(stats_text(snap, a.json))
        return

    if a.cmd == 'verify':
        with open(a.json, encoding='utf-8') as f:
            snap = json.load(f)
        r = verify_totals(snap, a.xlsx)
        tot = r['재계산일치'] + r['재계산불일치']
        pct = r['재계산일치'] / tot * 100 if tot else 0
        print(f"[역검증] {os.path.basename(a.json)}")
        print(f"  JSON 만으로 총점 재계산: {r['재계산일치']}/{tot} 일치 ({pct:.1f}%)")
        if r['불일치목록']:
            print(f"  불일치 예: {r['불일치목록']}")
        if a.xlsx:
            stot = r['원본일치'] + r['원본불일치']
            spct = r['원본일치'] / stot * 100 if stot else 0
            print(f"  원본 xlsx 총점 대조: {r['원본일치']}/{stot} 일치 ({spct:.1f}%)")
            if r['원본불일치목록']:
                print(f"  불일치 예: {r['원본불일치목록']}")
        if r['재계산불일치'] or r['원본불일치']:
            sys.exit(1)
        return

    if a.cmd == 'manifest':
        man = build_manifest(a.kr, a.us, a.runner, a.memo)
        size = write_json(a.out, man)
        print(f'저장: {a.out} ({size:,} 바이트)')
        print(json.dumps(man, ensure_ascii=False, indent=2))
        return

    if a.cmd == 'strategies':
        data = build_strategies(a.src)
        size = write_json(a.out, data)
        print(f"저장: {a.out} ({size:,} 바이트) — 전략 {len(data['전략'])}개, "
              f"필터설명 {len(data['필터설명'])}개")
        return


if __name__ == '__main__':
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    main()
