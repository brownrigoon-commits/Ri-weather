# -*- coding: utf-8 -*-
"""골프장 DB(golfdb.js)를 엑셀로 내보내기 — 확인용
한국어명(k) 또는 한국 구장(c=KR)을 포함해 국가별 시트로 정리.
출력: Ri-weather/골프장DB.xlsx
"""
import json, os, re, sys
sys.stdout.reconfigure(encoding="utf-8")
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("openpyxl 필요: pip install openpyxl"); sys.exit(1)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
t = open(os.path.join(ROOT, "js", "golfdb.js"), encoding="utf-8").read()
m = re.search(r"const GOLF_DB = (\[.*\]);", t, re.S)
arr = json.loads(m.group(1))

CO = {"KR": "한국", "JP": "일본", "CN": "중국"}
wb = Workbook()
wb.remove(wb.active)

hdr_font = Font(bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="2F6D4A")

def has_korean(s):
    return bool(re.search(r"[가-힣]", s or ""))

total = 0
for code, kor in CO.items():
    rows = []
    for g in arr:
        if g.get("c") != code:
            continue
        # 한국어가 포함된 것만: 한국명(k) 있거나 이름 자체가 한글
        name = g.get("n", "")
        kname = g.get("k", "")
        if code == "KR":
            display = name
        else:
            if not (has_korean(kname) or has_korean(name)):
                continue
            display = kname or name
        rows.append([display, name, g.get("a", ""), round(g.get("lat", 0), 5), round(g.get("lon", 0), 5)])
    rows.sort(key=lambda r: r[0])
    ws = wb.create_sheet(f"{kor}({len(rows)})")
    ws.append(["한국어 표기", "원문 이름", "별칭", "위도", "경도"])
    for c in ws[1]:
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append(r)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 11
    ws.column_dimensions["E"].width = 11
    ws.freeze_panes = "A2"
    total += len(rows)
    print(f"{kor}: {len(rows)}곳")

out = os.path.join(ROOT, "골프장DB.xlsx")
wb.save(out)
print(f"저장: {out} (총 {total}곳)")
