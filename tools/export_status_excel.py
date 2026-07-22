# -*- coding: utf-8 -*-
"""골프장DB.xlsx 재생성 — 홀별 공략 제작 현황 분류 포함
시트 구성
  ① 한국(전체)   : 골프DB 전 구장 + 수집/등록 상태 컬럼 (등급별 색상)
  ② 등록완료     : 앱에 홀별 공략이 들어간 구장
  ③ 제작가능     : 자료 확보돼 파이프라인만 돌리면 되는 구장
  ④ 부분수집     : 자료 일부만 (재수집·수동 필요)
  ⑤ 자료없음     : 홈페이지 미확보/홀맵 없음
  ⑥ 일본 / ⑦ 중국 : 참고용 목록
출력: Ri-weather/골프장DB.xlsx
"""
import glob, json, os, re, sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WF = os.path.join(ROOT, "coursedata", "workfiles")

def norm(s):
    return re.sub(r"(CC|GC|C\.C|G\.C|컨트리클럽|골프클럽|골프장|골프앤리조트|골프리조트|리조트|컨트리|클럽|\s|·|&|\(.*?\))", "", s or "", flags=re.I).lower()

# ── 데이터 로드 ───────────────────────────────────────────────
t = open(os.path.join(ROOT, "js", "golfdb.js"), encoding="utf-8").read()
DB = json.loads(re.search(r"const GOLF_DB = (\[.*\]);", t, re.S).group(1))

analysis = {}
p = os.path.join(WF, "registrable_analysis.json")
if os.path.exists(p):
    for r in json.load(open(p, encoding="utf-8")):
        analysis[norm(r["club"])] = r

report = {}
p = os.path.join(WF, "universal_build_report.json")
if os.path.exists(p):
    for r in json.load(open(p, encoding="utf-8")):
        report[norm(r["club"])] = r

registered = {}
for f in glob.glob(os.path.join(ROOT, "coursedata", "homepages", "*", "parsed.json")):
    d = json.load(open(f, encoding="utf-8"))
    registered[norm(d["course"])] = {
        "holes": sum(len(c["holes"]) for c in d["courses"]),
        "courses": ", ".join(c["name"] for c in d["courses"]),
        "tips": sum(1 for c in d["courses"] for h in c["holes"] if h.get("tip")),
        "src": d.get("sourceUrl", ""),
    }

missing_hp = {}
p = os.path.join(ROOT, "coursedata", "homepages_missing.json")
if os.path.exists(p):
    for k, v in json.load(open(p, encoding="utf-8")).items():
        if v:
            missing_hp[norm(k)] = v.get("url", "")

# ── 구장별 상태 판정 ──────────────────────────────────────────
STATUS = {
    "등록완료": ("✅ 등록완료", "C6EFCE"),
    "제작가능": ("🟢 제작가능", "D9F2D0"),
    "부분수집": ("🟡 부분수집", "FFF2CC"),
    "자료부족": ("🟠 자료부족", "FCE4D6"),
    "자료없음": ("⬜ 자료없음", "F2F2F2"),
}

def status_of(name):
    k = norm(name)
    if k in registered:
        r = registered[k]
        return "등록완료", f"{r['holes']}홀 ({r['courses']})", r["tips"], r["src"]
    rep = report.get(k)
    if rep and rep.get("ok"):
        return "제작가능", f"{rep['holes']}홀 파싱 성공 ({rep['parser']})", rep.get("tip", 0), ""
    an = analysis.get(k)
    if an:
        n = an["hole_imgs"]
        note = f"홀이미지 {n}장"
        if an.get("official_holes"):
            note += f" / 공식 {an['official_holes']}홀"
        if rep and rep.get("reason"):
            note += f" · {rep['reason'][:40]}"
        if an["grade"] in ("A", "B"):
            return "부분수집", note, an["tip_pages"], an.get("seed", "")
        if an["grade"] in ("C", "D"):
            return "자료부족", note, an["tip_pages"], an.get("seed", "")
        return "자료없음", note or "수집 실패", 0, an.get("seed", "")
    if k in missing_hp:
        return "자료없음", "홈페이지 주소만 확보(미수집)", 0, missing_hp[k]
    return "자료없음", "미수집", 0, ""

# ── 엑셀 작성 ─────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill("solid", fgColor="2F6D4A")

def write_sheet(ws, rows, cols, widths):
    ws.append(cols)
    for c in ws[1]:
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    for r in rows:
        ws.append(r)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"

kr = [g for g in DB if g.get("c") == "KR"]
rows_all, buckets = [], {k: [] for k in STATUS}
for g in sorted(kr, key=lambda x: x["n"]):
    st, note, tips, url = status_of(g["n"])
    row = [STATUS[st][0], g["n"], note, tips, url, round(g.get("lat", 0), 5), round(g.get("lon", 0), 5)]
    rows_all.append(row)
    buckets[st].append(row)

COLS = ["상태", "골프장", "상세", "공략TIP수", "출처/홈페이지", "위도", "경도"]
W = [13, 30, 46, 10, 44, 10, 10]

ws = wb.create_sheet(f"한국 전체({len(rows_all)})")
write_sheet(ws, rows_all, COLS, W)
for r in range(2, ws.max_row + 1):
    key = str(ws.cell(r, 1).value)
    for k, (label, color) in STATUS.items():
        if label == key:
            fill = PatternFill("solid", fgColor=color)
            for c in range(1, len(COLS) + 1):
                ws.cell(r, c).fill = fill
            break

for k in ["등록완료", "제작가능", "부분수집", "자료부족", "자료없음"]:
    b = buckets[k]
    ws = wb.create_sheet(f"{k}({len(b)})")
    write_sheet(ws, b, COLS, W)
    fill = PatternFill("solid", fgColor=STATUS[k][1])
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(COLS) + 1):
            ws.cell(r, c).fill = fill

def has_ko(s):
    return bool(re.search(r"[가-힣]", s or ""))

for code, kor in (("JP", "일본"), ("CN", "중국")):
    rows = []
    for g in DB:
        if g.get("c") != code:
            continue
        kn = g.get("k") or ""
        if not (has_ko(kn) or has_ko(g.get("n", ""))):
            continue
        rows.append([kn or g["n"], g["n"], g.get("a", ""), round(g.get("lat", 0), 5), round(g.get("lon", 0), 5)])
    rows.sort(key=lambda r: r[0])
    ws = wb.create_sheet(f"{kor}({len(rows)})")
    write_sheet(ws, rows, ["한국어 표기", "원문 이름", "별칭", "위도", "경도"], [30, 32, 16, 10, 10])

out = os.path.join(ROOT, "골프장DB.xlsx")
try:
    wb.save(out)
except PermissionError:
    out = os.path.join(ROOT, "골프장DB_현황.xlsx")   # 원본이 열려 있으면 별도 저장
    wb.save(out)
    print("※ 골프장DB.xlsx가 열려 있어 별도 파일로 저장했습니다.")
print("저장:", out)
for k in ["등록완료", "제작가능", "부분수집", "자료부족", "자료없음"]:
    print(f"  {STATUS[k][0]}: {len(buckets[k])}곳")
print(f"  한국 합계 {len(rows_all)}곳")
